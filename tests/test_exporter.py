"""Unit tests for Live Excel Manager and Exporter module."""

from pathlib import Path
import openpyxl
from scraper.config import OUTPUT_EXCEL_COLUMNS
from scraper.exporter import LiveExcelManager, export_sellers_to_excel


def test_live_excel_manager_immediate_save_and_update(tmp_path: Path):
    excel_path = tmp_path / "live_sellers.xlsx"
    manager = LiveExcelManager(output_path=excel_path)

    # 1. Initial Extraction (Status: ENRICHMENT_PENDING)
    initial_seller = {
        "seller_name": "IKAGIFOOTWEAR",
        "fulfillment_by": "IKAGIFOOTWEAR",
        "marketplace": "flipkart",
        "status": "ENRICHMENT_PENDING",
        "product_url": "https://www.flipkart.com/item/p/123",
        "category": "Footwear",
        "sub_category": "Men's Footwear",
        "product_rating": 3.9,
        "star_rating": None,
    }

    row1 = manager.write_or_update_seller(initial_seller)
    assert row1 == 1
    assert excel_path.exists()

    # Verify initial row in workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    assert ws.max_row == 2
    assert ws.cell(row=2, column=manager.header_col_map["Business Name"]).value == "IKAGIFOOTWEAR"
    assert ws.cell(row=2, column=manager.header_col_map["Status"]).value == "ENRICHMENT_PENDING"

    # 2. Another seller added
    seller_2 = {
        "seller_name": "LONDONSTEPS",
        "fulfillment_by": "LONDONSTEPS",
        "marketplace": "flipkart",
        "status": "ENRICHMENT_PENDING",
        "product_url": "https://www.flipkart.com/item/p/456",
        "category": "Footwear",
    }
    row2 = manager.write_or_update_seller(seller_2)
    assert row2 == 2

    # 3. Seller 1 Enrichment Complete -> Update existing row
    enriched_seller_1 = {
        "seller_name": "IKAGIFOOTWEAR",
        "gst_number": "27AAPFU0939F1ZV",
        "pan_number": "AAPFU0939F",
        "contact_number": "9876543210",
        "email": "contact@ikagi.in",
        "status": "VERIFIED",
        "website_url": "https://ikagi.in",
    }
    updated_row = manager.write_or_update_seller(enriched_seller_1)
    assert updated_row == 1  # Updated in-place on row 1!

    # Check that workbook still has only 2 data rows (no duplicates created)
    wb2 = openpyxl.load_workbook(excel_path)
    ws2 = wb2.active
    assert ws2.max_row == 3

    # Check updated fields
    assert ws2.cell(row=2, column=manager.header_col_map["GST Number"]).value == "27AAPFU0939F1ZV"
    assert ws2.cell(row=2, column=manager.header_col_map["Status"]).value == "VERIFIED"
    assert ws2.cell(row=2, column=manager.header_col_map["Email Address"]).value == "contact@ikagi.in"
    assert ws2.cell(row=2, column=manager.header_col_map["Website URL"]).value == "https://ikagi.in"


def test_export_sellers_to_excel_bulk(tmp_path: Path):
    output_xlsx = tmp_path / "flipkart_sellers_test.xlsx"

    sample_sellers = [
        {
            "seller_name": "Omnitech Retail",
            "business_model": "Proprietorship / Registered Business",
            "business_category": "Electronics > Mobiles",
            "owner_name": "Rajesh Sharma",
            "contact_number": "9876543210",
            "email": "contact@omnitech.in",
            "gst_number": "27AAPFU0939F1ZV",
            "pan_number": "AAPFU0939F",
            "fssai_number": None,
            "billing_address": "Plot 14, MIDC, Andheri East, Mumbai, Maharashtra 400069",
            "shipping_address": "Plot 14, MIDC, Andheri East, Mumbai, Maharashtra 400069",
            "city": "Mumbai",
            "state": "Maharashtra",
            "pincode": "400069",
            "country": "India",
            "website_url": "https://omnitech.in",
            "status": "VERIFIED",
            "source": ["company_website", "gst_portal"],
            "star_rating": 4.8,
        }
    ]

    exported_path = export_sellers_to_excel(sample_sellers, output_xlsx)
    assert exported_path.exists()

    wb = openpyxl.load_workbook(exported_path)
    ws = wb.active
    assert ws.max_row == 2
    headers = [cell.value for cell in ws[1]]
    assert headers == OUTPUT_EXCEL_COLUMNS


def test_live_excel_manager_verify_saved_row(tmp_path: Path):
    excel_path = tmp_path / "verify_test.xlsx"
    manager = LiveExcelManager(output_path=excel_path)

    seller = {
        "seller_name": "TRIPR",
        "gst_number": "33AAKCT0058N1Z3",
        "pan_number": "AAKCT0058N",
        "contact_number": "8031406054",
        "email": "queries@triprindia.com",
        "country": "India",
        "status": "VERIFIED",
        "website_url": "https://triprindia.com",
    }

    row_num = manager.write_or_update_seller(seller)
    assert row_num == 1

    # Verify real disk persistence
    verified, msg = manager.verify_saved_row(row_num, seller)
    assert verified is True
    assert "All fields verified" in msg

    # Verification fails on mismatch
    mismatch_seller = dict(seller)
    mismatch_seller["gst_number"] = "WRONG_GST"
    verified2, msg2 = manager.verify_saved_row(row_num, mismatch_seller)
    assert verified2 is False
    assert "gst number mismatch" in msg2.lower()


def test_live_excel_manager_schema_migration(tmp_path: Path):
    excel_path = tmp_path / "legacy_schema.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["business_model", "status", "star_rating"])
    ws.append(["Retail", "VERIFIED", 4.5])
    wb.save(excel_path)

    # Load with LiveExcelManager -> should migrate to exact 18 columns
    manager = LiveExcelManager(output_path=excel_path)
    assert list(manager.header_col_map.keys()) == OUTPUT_EXCEL_COLUMNS

    # Check that sheet was migrated on disk
    wb_migrated = openpyxl.load_workbook(excel_path)
    ws_migrated = wb_migrated.active
    headers = [cell.value for cell in ws_migrated[1]]
    assert headers == OUTPUT_EXCEL_COLUMNS

