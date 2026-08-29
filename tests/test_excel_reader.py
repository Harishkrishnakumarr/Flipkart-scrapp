"""Unit tests for Adaptive Excel Reader module."""

import pytest
from pathlib import Path
import openpyxl
from scraper.excel_reader import read_categories_from_excel


def test_read_categories_from_excel(tmp_path: Path):
    excel_path = tmp_path / "test_input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["category", "sub_category", "sub_sub_category", "sub_sub_subcategory"])
    ws.append(["Electronics", "Mobiles", "Smartphones", "Android Phones"])
    ws.append(["Fashion", "Footwear", "Men's Footwear", ""])
    ws.append([None, None, None, None])  # Empty row to skip
    wb.save(excel_path)

    tasks = read_categories_from_excel(excel_path, resume=False)
    assert len(tasks) == 2
    assert tasks[0]["query"] == "Electronics Mobiles Smartphones Android Phones"
    assert tasks[1]["query"] == "Fashion Footwear Men's Footwear"
    assert tasks[0]["row_index"] == 2
    assert tasks[1]["row_index"] == 3


def test_read_direct_product_urls_from_excel(tmp_path: Path):
    excel_path = tmp_path / "test_urls.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["product_url", "category"])
    ws.append(["https://www.flipkart.com/sample-product/p/itm123", "Footwear"])
    wb.save(excel_path)

    tasks = read_categories_from_excel(excel_path, resume=False)
    assert len(tasks) == 1
    assert tasks[0]["product_url"] == "https://www.flipkart.com/sample-product/p/itm123"
    assert tasks[0]["category_dict"]["category"] == "Footwear"


def test_read_empty_header_raises_error(tmp_path: Path):
    excel_path = tmp_path / "empty_input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None, None])
    wb.save(excel_path)

    with pytest.raises(ValueError, match="empty header row"):
        read_categories_from_excel(excel_path, resume=False)


def test_extract_google_sheet_details():
    from scraper.excel_reader import extract_google_sheet_details

    url1 = "https://docs.google.com/spreadsheets/d/1CLux0vBFcNH3MR03x2afq4GaJ3zhIPJ_Y-xq4lKXghk/edit?gid=899569498#gid=899569498"
    res1 = extract_google_sheet_details(url1)
    assert res1["spreadsheet_id"] == "1CLux0vBFcNH3MR03x2afq4GaJ3zhIPJ_Y-xq4lKXghk"
    assert res1["gid"] == "899569498"

    raw_id = "1CLux0vBFcNH3MR03x2afq4GaJ3zhIPJ_Y-xq4lKXghk"
    res2 = extract_google_sheet_details(raw_id)
    assert res2["spreadsheet_id"] == raw_id
    assert res2["gid"] == "0"

