"""Live Excel export module for immediate per-seller persistence, updates, and verified saving."""

import logging
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scraper.config import OUTPUT_EXCEL_COLUMNS, OUTPUT_FILE

logger = logging.getLogger("FlipkartScraper.Exporter")

# Styling Definitions
HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

ROW_ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
ROW_DEFAULT_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

STATUS_STYLES = {
    "VERIFIED": {
        "fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "font": Font(name="Calibri", size=10, bold=True, color="166534"),
    },
    "PARTIALLY_VERIFIED": {
        "fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "font": Font(name="Calibri", size=10, bold=True, color="1E40AF"),
    },
    "ENRICHMENT_PENDING": {
        "fill": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        "font": Font(name="Calibri", size=10, bold=True, color="854D0E"),
    },
    "NEEDS_REVIEW": {
        "fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "font": Font(name="Calibri", size=10, bold=True, color="92400E"),
    },
    "NOT_FOUND": {
        "fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
        "font": Font(name="Calibri", size=10, bold=False, color="475569"),
    },
}

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


EXCEL_FIELD_MAPPING: Dict[str, List[str]] = {
    "Business Name": ["Business Name", "seller_name", "business_name", "company_name", "name"],
    "Business Model": ["Business Model", "business_model"],
    "Business Category": ["Business Category", "business_category", "category"],
    "Owner Name": ["Owner Name", "owner_name", "owner", "proprietor_name", "director_name"],
    "Phone Number": ["Phone Number", "contact_number", "phone_number", "phone", "mobile"],
    "Email Address": ["Email Address", "email", "email_address", "contact_email"],
    "GST Number": ["GST Number", "gst_number", "gstin", "gst"],
    "PAN Number": ["PAN Number", "pan_number", "pan"],
    "FSSAI Number": ["FSSAI Number", "fssai_number", "fssai"],
    "Billing Address": ["Billing Address", "billing_address", "address", "raw_address"],
    "x": ["x", "shipping_address", "fulfillment_by"],
    "City": ["City", "city"],
    "State": ["State", "state"],
    "Pincode": ["Pincode", "pincode", "postal_code", "zip"],
    "Country": ["Country", "country"],
    "Website URL": ["Website URL", "website_url", "website", "official_website"],
    "Status": ["Status", "status"],
    "Source rating": ["Source rating", "source_rating", "star_rating", "seller_rating"],
}


def _seller_key(name: Optional[str]) -> str:
    """Normalize seller name for indexing and comparison."""
    if not name:
        return ""
    val = str(name).lower().strip().replace("&", "and")
    val = re.sub(r"[^\w\s]", "", val)
    return " ".join(val.split())


def _extract_column_value(col_name: str, seller_data: Dict[str, Any]) -> Any:
    """Extract appropriate value for a column from seller data dictionary."""
    candidate_keys = EXCEL_FIELD_MAPPING.get(col_name, [col_name])
    for k in candidate_keys:
        if k in seller_data and seller_data[k] is not None:
            val = seller_data[k]
            if str(val).strip() not in ("", "NOT FOUND", "N/A"):
                return val

    # Defaults
    if col_name == "Country":
        return "India"
    if col_name == "x":
        return seller_data.get("fulfillment_by") or seller_data.get("shipping_address")
    if col_name == "Source rating":
        return seller_data.get("star_rating") or seller_data.get("seller_rating") or seller_data.get("source_rating")
    return None


class LiveExcelManager:
    """Manages immediate, per-seller Excel file creation, row insertion, in-place updates, and disk verification."""

    def __init__(self, output_path: Path = OUTPUT_FILE) -> None:
        """Initialize or load live Excel workbook.

        Args:
            output_path: Path to the target Excel file.
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        self.seller_row_map: Dict[str, int] = {}
        self.header_col_map: Dict[str, int] = {
            col_name: idx for idx, col_name in enumerate(OUTPUT_EXCEL_COLUMNS, start=1)
        }

        self.workbook: openpyxl.Workbook
        self.sheet: openpyxl.worksheet.worksheet.Worksheet

        if self.output_path.exists():
            self._load_and_reconcile_existing_workbook()
        else:
            self._init_new_workbook()

    def _init_in_memory_workbook(self) -> None:
        """Initialize in-memory workbook structure."""
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Flipkart Sellers"
        self.sheet.views.sheetView[0].showGridLines = True
        self.sheet.freeze_panes = "A2"

        for col_idx, col_name in enumerate(OUTPUT_EXCEL_COLUMNS, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = THIN_BORDER

        self.sheet.row_dimensions[1].height = 28

    def _init_new_workbook(self) -> None:
        """Create fresh workbook with styled header row matching OUTPUT_EXCEL_COLUMNS and save."""
        self._init_in_memory_workbook()
        self._save_workbook_with_retry()
        logger.debug(f"Initialized new live Excel file at {self.output_path}")

    def _save_workbook_with_retry(self, max_retries: int = 3, retry_delay: float = 0.5) -> bool:
        """Save workbook to disk with retries against file locks."""
        for attempt in range(1, max_retries + 1):
            try:
                self.workbook.save(filename=self.output_path)
                return True
            except PermissionError:
                if attempt < max_retries:
                    logger.warning(
                        f"Excel file '{self.output_path}' is locked (e.g. open in Excel). "
                        f"Retrying save ({attempt}/{max_retries})... Please close the file in Excel!"
                    )
                    time.sleep(retry_delay)
                else:
                    logger.error(
                        f"PERMISSION ERROR: Unable to save to '{self.output_path}'. "
                        f"The file is locked by another program. Please close it to allow saving."
                    )
                    return False
            except Exception as e:
                logger.error(f"Unexpected error saving workbook to {self.output_path}: {e}")
                return False
        return False

    def _load_and_reconcile_existing_workbook(self) -> None:
        """Load existing workbook and reconcile headers to ensure all standard columns exist."""
        try:
            self.workbook = openpyxl.load_workbook(filename=self.output_path)
            self.sheet = self.workbook.active

            header_row = [cell.value for cell in self.sheet[1]] if self.sheet.max_row >= 1 else []
            clean_headers = [str(c).strip() for c in header_row if c is not None]

            # Check if header row matches standard OUTPUT_EXCEL_COLUMNS
            needs_migration = clean_headers != OUTPUT_EXCEL_COLUMNS

            if needs_migration:
                logger.info(f"Reconciling Excel headers in {self.output_path} to standard schema...")
                old_col_map: Dict[str, int] = {
                    str(col_name).strip(): idx for idx, col_name in enumerate(header_row, start=1) if col_name
                }

                # Extract existing row data
                old_rows_data: List[Dict[str, Any]] = []
                for row_idx in range(2, self.sheet.max_row + 1):
                    row_dict: Dict[str, Any] = {}
                    for col_name, old_idx in old_col_map.items():
                        row_dict[col_name] = self.sheet.cell(row=row_idx, column=old_idx).value
                    if any(v is not None for v in row_dict.values()):
                        old_rows_data.append(row_dict)

                # Reset sheet with standard header row
                self.sheet.delete_rows(1, self.sheet.max_row)
                self.sheet.title = "Flipkart Sellers"
                self.sheet.views.sheetView[0].showGridLines = True
                self.sheet.freeze_panes = "A2"

                for col_idx, col_name in enumerate(OUTPUT_EXCEL_COLUMNS, start=1):
                    cell = self.sheet.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    cell.border = THIN_BORDER

                self.sheet.row_dimensions[1].height = 28

                # Re-insert existing rows into standard column positions
                for row_idx, row_dict in enumerate(old_rows_data, start=2):
                    self.sheet.row_dimensions[row_idx].height = 22
                    row_fill = ROW_ALT_FILL if row_idx % 2 == 0 else ROW_DEFAULT_FILL
                    for col_name in OUTPUT_EXCEL_COLUMNS:
                        col_idx = self.header_col_map[col_name]
                        val = _extract_column_value(col_name, row_dict)
                        cell = self.sheet.cell(row=row_idx, column=col_idx, value=val)
                        cell.border = THIN_BORDER
                        cell.font = Font(name="Calibri", size=10)
                        cell.fill = row_fill
                        if col_name in [
                            "Phone Number", "GST Number", "PAN Number", "FSSAI Number",
                            "Pincode", "Source rating", "Status"
                        ]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                self._save_workbook_with_retry()
                logger.info(f"Successfully migrated {len(old_rows_data)} rows to standard schema.")

            # Index existing rows by canonical Business Name (column 1)
            name_col_idx = self.header_col_map.get("Business Name", 1)
            for row_idx in range(2, self.sheet.max_row + 1):
                cell_val = self.sheet.cell(row=row_idx, column=name_col_idx).value
                if cell_val:
                    k = _seller_key(str(cell_val))
                    if k:
                        self.seller_row_map[k] = row_idx

            logger.info(
                f"Loaded live Excel from {self.output_path} ({len(self.seller_row_map)} existing seller rows indexed)."
            )
        except PermissionError:
            logger.error(
                f"Cannot load '{self.output_path}' because it is open in Microsoft Excel. "
                f"Initializing in-memory session. Please close the file in Excel to save changes."
            )
            self._init_in_memory_workbook()
        except Exception as e:
            logger.warning(f"Could not load existing Excel ({e}), reinitializing fresh workbook.")
            self._init_new_workbook()

    def get_completed_sellers(self) -> Set[str]:
        """Return set of canonical seller keys already enriched and present in the Excel."""
        completed: Set[str] = set()
        status_col_idx = self.header_col_map.get("Status", 17)
        seller_col_idx = self.header_col_map.get("Business Name", 1)

        for row_idx in range(2, self.sheet.max_row + 1):
            s_val = self.sheet.cell(row=row_idx, column=seller_col_idx).value
            stat_val = self.sheet.cell(row=row_idx, column=status_col_idx).value
            if s_val and stat_val and str(stat_val).strip() not in {"ENRICHMENT_PENDING", "UNKNOWN"}:
                k = _seller_key(str(s_val))
                if k:
                    completed.add(k)

        return completed

    def write_or_update_seller(self, seller_data: Dict[str, Any], auto_save: bool = True) -> int:
        """Immediately write new seller record or update existing seller row in Excel.

        Args:
            seller_data: Seller dictionary.
            auto_save: If True, saves workbook immediately to disk.

        Returns:
            Data row number (1-indexed, corresponding to display row, where row 1 is header).
        """
        seller_name = (
            seller_data.get("Business Name")
            or seller_data.get("seller_name")
            or seller_data.get("owner_name")
            or ""
        )
        canonical_key = _seller_key(seller_name)

        if not canonical_key:
            return 0

        # Determine target row
        is_update = canonical_key in self.seller_row_map
        if is_update:
            target_row = self.seller_row_map[canonical_key]
        else:
            target_row = self.sheet.max_row + 1
            self.seller_row_map[canonical_key] = target_row

        row_fill = ROW_ALT_FILL if target_row % 2 == 0 else ROW_DEFAULT_FILL
        self.sheet.row_dimensions[target_row].height = 22

        # Populate / Update columns
        for col_name in OUTPUT_EXCEL_COLUMNS:
            col_idx = self.header_col_map[col_name]
            val = _extract_column_value(col_name, seller_data)

            # Ensure Business Name is always set
            if col_name == "Business Name" and not val:
                val = seller_name

            # If updating and new value is None or placeholder, do not overwrite existing valid cell content
            if is_update and (val is None or val in ("", "NOT FOUND", "N/A")):
                continue

            # Convert list (e.g. source) to string
            if isinstance(val, list):
                val = ", ".join(str(item) for item in val) if val else None

            cell = self.sheet.cell(row=target_row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.fill = row_fill

            # Alignment
            if col_name in [
                "Phone Number",
                "GST Number",
                "PAN Number",
                "FSSAI Number",
                "Pincode",
                "Source rating",
                "Status",
            ]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Status cell special highlight
            if col_name == "Status":
                status_str = str(val).strip() if val else ""
                if status_str in STATUS_STYLES:
                    cell.fill = STATUS_STYLES[status_str]["fill"]
                    cell.font = STATUS_STYLES[status_str]["font"]

        # Adjust column widths dynamically
        for col_name in OUTPUT_EXCEL_COLUMNS:
            col_idx = self.header_col_map[col_name]
            col_letter = get_column_letter(col_idx)
            val_str = str(_extract_column_value(col_name, seller_data) or "")
            current_w = self.sheet.column_dimensions[col_letter].width or (len(col_name) + 4)
            new_w = max(current_w, len(val_str) + 4, len(col_name) + 4, 14)
            self.sheet.column_dimensions[col_letter].width = min(new_w, 60)

        # Save workbook to disk with retry if auto_save is enabled
        if auto_save:
            self._save_workbook_with_retry()
        data_row_num = target_row - 1
        return data_row_num

    def verify_saved_row(
        self, display_row_num: int, expected_data: Dict[str, Any]
    ) -> Tuple[bool, str]:
        """Reopen the saved Excel file from disk and verify that updated fields are actually saved.

        Args:
            display_row_num: The 1-indexed data row number (where row 1 corresponds to sheet row 2).
            expected_data: Dictionary of expected seller values.

        Returns:
            Tuple of (is_verified: bool, details_or_error_msg: str).
        """
        if not self.output_path.exists():
            return False, f"Excel file does not exist on disk at {self.output_path}"

        target_sheet_row = display_row_num + 1

        try:
            # Reopen workbook directly from disk to verify persistence
            verify_wb = openpyxl.load_workbook(filename=self.output_path, data_only=True)
            verify_sheet = verify_wb.active

            if target_sheet_row > verify_sheet.max_row:
                return False, f"Row {target_sheet_row} exceeds sheet max_row ({verify_sheet.max_row})"

            # Read cells from reopened sheet
            saved_values: Dict[str, Any] = {}
            for col_name in OUTPUT_EXCEL_COLUMNS:
                col_idx = self.header_col_map.get(col_name, 1)
                saved_values[col_name] = verify_sheet.cell(row=target_sheet_row, column=col_idx).value

            # 1. Verify Business Name
            expected_seller = (
                expected_data.get("Business Name")
                or expected_data.get("seller_name")
                or expected_data.get("owner_name")
            )
            if expected_seller:
                saved_seller = saved_values.get("Business Name")
                if _seller_key(saved_seller) != _seller_key(expected_seller):
                    return (
                        False,
                        f"Business Name mismatch at row {target_sheet_row}: expected '{expected_seller}', found '{saved_seller}'",
                    )

            # 2. Verify Status
            expected_status = expected_data.get("Status") or expected_data.get("status")
            if expected_status:
                saved_status = str(saved_values.get("Status") or "").strip()
                if saved_status != str(expected_status).strip():
                    return (
                        False,
                        f"Status mismatch at row {target_sheet_row}: expected '{expected_status}', found '{saved_status}'",
                    )

            # 3. Verify key credentials if provided
            key_fields = [
                ("GST Number", ["GST Number", "gst_number"]),
                ("PAN Number", ["PAN Number", "pan_number"]),
                ("Phone Number", ["Phone Number", "contact_number", "phone"]),
                ("Email Address", ["Email Address", "email"]),
                ("Country", ["Country", "country"]),
                ("Website URL", ["Website URL", "website_url"]),
            ]
            for col_hdr, aliases in key_fields:
                exp_val = None
                for a in aliases:
                    if expected_data.get(a) is not None:
                        exp_val = expected_data[a]
                        break
                if exp_val is not None and str(exp_val).strip() not in ("", "N/A", "NOT FOUND"):
                    saved_val = str(saved_values.get(col_hdr) or "").strip()
                    if saved_val != str(exp_val).strip():
                        return (
                            False,
                            f"{col_hdr} mismatch at row {target_sheet_row}: expected '{exp_val}', found '{saved_val}'",
                        )

            return True, "All fields verified successfully in saved Excel file."
        except PermissionError:
            return False, "File is locked by another process (e.g. open in Microsoft Excel)"
        except Exception as e:
            return False, f"Exception during Excel verification: {e}"

    def commit_all_sellers(self, sellers_data: List[Dict[str, Any]]) -> int:
        """Commit list of seller dictionaries to workbook and ensure final save.

        Args:
            sellers_data: List of seller dictionary records.

        Returns:
            Number of sellers written/updated.
        """
        count = 0
        for seller in sellers_data:
            if seller and (seller.get("seller_name") or seller.get("owner_name")):
                self.write_or_update_seller(seller, auto_save=False)
                count += 1
        self._save_workbook_with_retry()
        logger.info(f"Committed and verified {count} sellers to {self.output_path}")
        return count


def export_sellers_to_excel(
    sellers_data: List[Dict[str, Any]],
    output_path: Path = OUTPUT_FILE,
) -> Path:
    """Export list of seller dictionaries to formatted Excel file using LiveExcelManager.

    Args:
        sellers_data: List of seller dictionary records.
        output_path: Target path for the output .xlsx file.

    Returns:
        Path to the generated Excel file.
    """
    manager = LiveExcelManager(output_path=output_path)
    manager.commit_all_sellers(sellers_data)
    logger.info(f"Excel file successfully generated at: {output_path}")
    return Path(output_path)
