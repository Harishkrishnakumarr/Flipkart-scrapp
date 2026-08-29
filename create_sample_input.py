"""Helper script to generate the sample input.xlsx file."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

def generate_sample_input_excel(file_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CategoryHierarchy"
    ws.views.sheetView[0].showGridLines = True

    headers = ["category", "sub_category", "sub_sub_category", "sub_sub_subcategory"]
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 26

    # Sample realistic Flipkart hierarchy rows
    rows = [
        ["Electronics", "Mobiles", "Smartphones", "Android Phones"],
        ["Electronics", "Audio", "Headphones", "Wireless Earbuds"],
        ["Electronics", "Laptops", "Gaming Laptops", "Core i7"],
        ["Home & Kitchen", "Kitchen Appliances", "Small Appliances", "Air Fryers"],
        ["Fashion", "Footwear", "Men's Footwear", "Running Shoes"],
        ["Beauty & Personal Care", "Skin Care", "Face Care", "Face Serums"],
        ["Sports & Fitness", "Fitness Equipment", "Cardio", "Treadmills"],
        ["Books", "Educational", "Competitive Exams", "Engineering"],
    ]

    for row in rows:
        ws.append(row)

    # Adjust widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 20)

    wb.save(file_path)
    print(f"Sample input Excel created at: {file_path}")

if __name__ == "__main__":
    generate_sample_input_excel(Path("input.xlsx"))
